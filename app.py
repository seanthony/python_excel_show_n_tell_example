from file_stitcher import get_tabular_data
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from json import load

BRAND = {
    "white": "ffffff",
    "navy": "081e32",
    "teal": "1996aa",
    "gray": "8e8884",
    "yellow": "ffe000",
    "red": "dc5756"
}


def get_chr(i):
    s = ''
    if i // 26 > 0:
        s = chr(65 + (i // 26))
    s += chr(65 + (i % 26))
    return s


def get_ranges(tabular_data):
    header_cells = [
        "I learned new skills or strategies today",
        "I was able to connect with someone new today",
        "This session met my needs as learner",
        "I came away from this session with a useful resource"
    ]

    ranges = []
    for i in range(len(tabular_data[0])):
        header = tabular_data[0][i]
        if header in header_cells:
            ranges.append(get_chr(i))

    return ranges


def format_data(tabular_data):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Survey Data"
    sheet.sheet_properties.tabColor = BRAND.get("teal")

    num_calculated_values = 5

    for _ in range(num_calculated_values):
        sheet.append([])

    for row in tabular_data:
        sheet.append(row)

    last_row = len(tabular_data) + num_calculated_values

    ranges = get_ranges(tabular_data)
    for r in ranges:
        formula_range = f'{r}7:{r}{last_row}'
        sheet[f'{r}1'] = f'={r}2 - {r}3'
        sheet[f'{r}2'] = f'=COUNTIF({formula_range}, ">5")/COUNT({formula_range})'
        sheet[f'{r}3'] = f'=COUNTIF({formula_range}, "<5")/COUNT({formula_range})'
        for i in range(1, 4):
            sheet[f'{r}{i}'].number_format = '0%'

        sheet[f'{r}4'] = '=CONCATENATE("(", ' + ', ", ", '.join(
            [f'COUNTIF({formula_range}, {i})' for i in range(1, 8)]) + ', ")")'

        sheet[f'{r}5'] = f'=AVERAGE({formula_range})'
        sheet[f'{r}5'].number_format = '0.00'

        for i in range(1, 6):
            sheet[f'{r}{i}'].alignment = Alignment(horizontal='center')

        sheet.column_dimensions[r].width = 20

    sheet.freeze_panes = 'D7'
    sheet['C1'] = "Net Strength >"
    sheet['C2'] = "% Strength >"
    sheet['C3'] = "% Weak >"
    sheet['C4'] = "Distribution >"
    sheet['C5'] = "Mean >"
    for i in range(1, 6):
        sheet[f'C{i}'].alignment = Alignment(horizontal='right')

    side = Side(style='thin', color=BRAND.get("navy"))
    thin_border = Border(left=side,
                         right=side,
                         top=side,
                         bottom=side)

    for row in sheet:
        for cell in row:
            cell.font = Font(name="Calibri")
            if cell.value:
                cell.border = thin_border

    for r in ranges:
        for i in range(2, 5):
            size = 10
            if i > 3:
                size = 8
            sheet[f'{r}{i}'].font = Font(
                name="Calibri", size=size, color=BRAND.get("gray"))

    data_range = 'A6:{}{}'.format(
        get_chr(len(tabular_data[0]) - 1), 6 + len(tabular_data))
    sheet.auto_filter.ref = data_range

    font = Font(name="Calibri", color=BRAND.get("white"))
    font_bold = Font(name="Calibri", bold="True", color=BRAND.get("white"))
    side = Side(border_style=None)
    no_border = Border(left=side, right=side, top=side, bottom=side)
    for i in range(len(tabular_data[0])):
        c = get_chr(i)
        cell = sheet[f'{c}6']
        cell.fill = PatternFill(fgColor=BRAND.get(
            'navy'), fill_type="solid", patternType="solid")
        cell.font = font_bold

        if c in ranges:
            continue

        for j in range(1, 6):
            cell = sheet[f'{c}{j}']
            cell.fill = PatternFill(fgColor=BRAND.get(
                'teal'), fill_type="solid", patternType="solid")
            cell.border = no_border
            cell.font = font

    return wb


def build_cohort_reports(tabular_data, folder, filename, cohort_indicies):
    for index in cohort_indicies:
        values = list(set(map(lambda r: r[index], tabular_data[1:])))
        for value in values:
            data_subset = [tabular_data[0]] + \
                list(filter(lambda r: r[index] == value, tabular_data[1:]))
            wb = format_data(data_subset)
            new_filename = '_'.join(
                f'{tabular_data[0][index]}_{value}_{filename}'.split())
            for ch in "'\"-:":
                new_filename = new_filename.replace(ch, '')

            filepath = f'{folder}Cohorts/{new_filename}'
            wb.save(filepath)


def main():
    with open('./settings.json') as file:
        settings = load(file)

    decision = input(
        'Overwrite existing file (Y for yes)?\n>>>').lower().strip() == 'y'

    tabular_data = get_tabular_data(overwrite=decision)
    wb = format_data(tabular_data)
    filename = "Formatted_Survey_Data.xlsx"
    wb.save(filename)  # saves local version
    wb.save(settings.get('destination folder') + "Formatted_Survey_Data.xlsx")

    build_cohort_reports(tabular_data, settings.get(
        'destination folder'), filename, [3, 4, 5, 8])


if __name__ == "__main__":
    main()
