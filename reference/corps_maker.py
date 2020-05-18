from random import randint, choice, sample, shuffle
from openpyxl import Workbook, load_workbook


def load_names(filepath):
    with open(filepath) as file:
        raw_names = file.readlines()

    return list(set(map(lambda name: name.strip(), raw_names)))


def generate_coach(names, number):
    coaches = set()
    while len(coaches) < number:
        name = f'{choice(names)} {choice(names)}'
        coaches.add(name)
    return coaches


def create_cms(names, coaches, grade_levels, corps_size):
    cms = [['PID', 'First', 'Last', 'Year', 'Grade Level', 'Coach']]
    pid = 30000
    while len(cms) <= corps_size:
        first = choice(names)
        last = choice(names)
        if randint(0, 100) < 10:
            last += '-' + choice(names)
        elif randint(0, 100) < 6:
            last += choice([', Jr.', ', II', ', III'])
        year = choice([2019, 2020])
        grade = choice(grade_levels)
        coach = sample(coaches, 1)[0]
        pid += randint(2, 10)
        row = [pid, first, last, year, grade, coach]
        cms.append(row)
    return cms


def create_cm_sheet(cms):
    wb = Workbook()
    sheet = wb["Sheet"]
    sheet.title = 'Corps Members'
    for row in cms:
        sheet.append(row)
    wb.save('./reference/cms.xlsx')


def make_cms():
    corps_size = randint(90, 140)
    cohorts = 6
    grade_levels = ['Lower el', 'Upper el', 'Middle', "High School"]
    names = load_names('./reference/names.txt')
    coaches = generate_coach(names, cohorts)
    cms = create_cms(names, coaches, grade_levels, corps_size)
    create_cm_sheet(cms)
    return cms


def load_cms():
    wb = load_workbook('./referemce/cms.xlsx')
    cms = [[cell.value for cell in row] for row in wb['Corps Members']]


def make_list_dict(cm_table):
    header, values = cm_table[0], cm_table[1:]
    cms = [{header[i]: row[i] for i in range(len(header))} for row in values]
    return cms


def make_survey_data(cms):
    shuffle(cms)
    n = len(cms)
    subset = sample(cms, randint(round(n * .7), n))
    responses = ['Strongly Disagree', 'Disagree', 'Somewhat Disagree',
                 'Neutral', 'Somewhat Agree', 'Agree', 'Strongly Agree']

    questions = ['I learned new skills or strategies today', 'I was able to connect with someone new today',
                 'Session Attended', 'This session met my needs as learner', 'I came away from this session with a useful resource']
    sessions = ["Basics of Blue Heeler Training", "Sean's Motorcycle Repair 101",
                "Using openpyxl to Create Reports", 'Speed Shoe Tying']

    data = [['Name', 'PID'] + questions]
    for cm in subset:
        name = '{} {}'.format(cm.get('First'), cm.get('Last'))
        row = [name, cm.get('PID')]
        for question in questions:
            if question == 'Session Attended':
                row.append(choice(sessions))
            else:
                row.append(choice(responses))
        data.append(row)

    wb = Workbook()
    sheet = wb["Sheet"]
    sheet.title = 'Raw Data'
    for row in data:
        sheet.append(row)
    wb.save('./reference/event_survey_data.xlsx')


def main():
    should_make_cms = True

    if should_make_cms:
        cm_table = make_cms()
    else:
        cm_table = load_cms()

    cms = make_list_dict(cm_table)
    make_survey_data(cms)


if __name__ == "__main__":
    main()
